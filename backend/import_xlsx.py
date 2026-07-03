# -*- coding: utf-8 -*-
"""
把 Excel 题库导入 SQLite。

数据源：data/招标代理2025考试题库.xlsx
- 3 个 Sheet：单选题 / 多选题 / 判断题
- 表头在第 2 行（第 1 行是标题）
- 判断题答案为 √ / ×
- 单选答案 A/B/C/D，多选为 AB/ABC/ABCD 等

归一化策略：
- options：收集非空选项，存 JSON 数组
- answer_norm：
    单选/多选 → 正确选项的索引数组，如 [0] 或 [0,1,2,3]
    判断题   → [true]（对）/ [false]（错）；前端交互时用 0 表示对、1 表示错（见 _judge）

用法：
    cd 招标代理题库_后端版
    python -m backend.import_xlsx            # 默认 data/招标代理2025考试题库.xlsx
    python -m backend.import_xlsx --reset    # 先清空 questions 再导入
"""
import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

# 让脚本既能被 -m 调用，也能直接 python backend/import_xlsx.py
ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.database import SessionLocal, init_db  # noqa: E402
from backend.models import Question  # noqa: E402

LETTER_TO_IDX = {ch: i for i, ch in enumerate("ABCDEFGH")}

SHEET_CONFIG = [
    ("单选题", "single"),
    ("多选题", "multi"),
    ("判断题", "judge"),
]


def _clean(cell):
    if cell is None:
        return ""
    return str(cell).strip()


def _norm_single_multi(letters: str):
    """A / ABCD -> [0] / [0,1,2,3]"""
    letters = (letters or "").upper().replace(" ", "")
    return [LETTER_TO_IDX[ch] for ch in letters if ch in LETTER_TO_IDX]


def _norm_judge(raw: str):
    """√/×/对/错/正确/错误/T/F -> [true]/[false]"""
    s = (raw or "").strip()
    if s in ("√", "对", "正确", "T", "true", "True"):
        return [True]
    if s in ("×", "错", "错误", "F", "false", "False"):
        return [False]
    return []


def parse_sheet(ws, qtype: str, type_label: str):
    """从某个 sheet 解析出 Question 字典列表。type_label 用于 qid 前缀（单选/多选/判断）。"""
    rows = list(ws.iter_rows(values_only=True))
    # 表头在第 2 行（idx=1），数据从 idx=2 起
    data = [r for r in rows[2:] if any(c is not None and _clean(c) for c in r)]
    items = []
    for r in data:
        qid = _clean(r[0])
        stem = _clean(r[1])
        if not stem:
            continue

        if qtype == "judge":
            options = ["对", "错"]
            answer_raw = _clean(r[2])
            answer_norm = _norm_judge(answer_raw)
            answer_disp = "对" if answer_norm == [True] else ("错" if answer_norm == [False] else answer_raw)
        else:
            # 选项列：答案A(2) 答案B(3) 答案C(4) 答案D(5)
            options = []
            for col in (2, 3, 4, 5):
                if col < len(r):
                    v = _clean(r[col])
                    if v:
                        options.append(v)
            answer_raw = _clean(r[6]) if len(r) > 6 else ""
            answer_norm = _norm_single_multi(answer_raw)
            answer_disp = answer_raw.upper()

        items.append({
            "qid": f"{type_label}-{qid}",
            "type": qtype,
            "stem": stem,
            "options": options,
            "answer": answer_disp,
            "answer_norm": answer_norm,
        })
    return items


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(ROOT / "data" / "招标代理2025考试题库.xlsx"))
    ap.add_argument("--reset", action="store_true", help="导入前清空 questions 表")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"[ERR] 找不到 Excel：{args.xlsx}")
        sys.exit(1)

    init_db()
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    all_items = []
    for sheet_name, qtype in SHEET_CONFIG:
        if sheet_name not in wb.sheetnames:
            print(f"[WARN] 缺少 sheet：{sheet_name}")
            continue
        type_label = {"single": "单选", "multi": "多选", "judge": "判断"}[qtype]
        items = parse_sheet(wb[sheet_name], qtype, type_label)
        print(f"  {sheet_name}: 解析到 {len(items)} 题")
        all_items.extend(items)

    db: Session = SessionLocal()
    try:
        if args.reset:
            deleted = db.query(Question).delete()
            print(f"[RESET] 清空 questions 表，删除 {deleted} 条")
        existing = db.query(Question).count()
        if existing and not args.reset:
            print(f"[INFO] questions 表已有 {existing} 条，继续追加。（用 --reset 可清空重导）")

        added = 0
        for it in all_items:
            db.add(Question(
                qid=it["qid"],
                type=it["type"],
                stem=it["stem"],
                options=json.dumps(it["options"], ensure_ascii=False),
                answer=it["answer"],
                answer_norm=json.dumps(it["answer_norm"], ensure_ascii=False),
            ))
            added += 1
        db.commit()
        total = db.query(Question).count()
        print(f"\n[DONE] 本次新增 {added} 条，当前 questions 表共 {total} 条")
    finally:
        db.close()


if __name__ == "__main__":
    main()
